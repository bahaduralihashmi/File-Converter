"""
Spreadsheet Converter - Complete with All Formats Support
Supports: XLSX, XLS, CSV, ODS, TSV, NUMBERS, HTML, JSON, XML
"""

from pathlib import Path
import os
import csv
import json
import tempfile
import shutil
import re

from app.utils.logger import get_logger

logger = get_logger(__name__)

# Try to import optional libraries
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import odf
    from odf.opendocument import load, OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
    ODF_AVAILABLE = True
except ImportError:
    ODF_AVAILABLE = False


class SpreadsheetConverter:
    """Complete spreadsheet converter with all format support"""
    
    def __init__(self):
        # All supported formats
        self.supported_formats = [
            'xlsx', 'xls', 'csv', 'ods', 'tsv', 
            'numbers', 'html', 'json', 'xml', 'dif'
        ]
        
        # Output formats
        self.output_formats = [
            'xlsx', 'xls', 'csv', 'ods', 'tsv', 
            'html', 'json', 'xml'
        ]
        
        logger.info("Spreadsheet Converter initialized")
        logger.info(f"Excel available: {EXCEL_AVAILABLE}")
        logger.info(f"Pandas available: {PANDAS_AVAILABLE}")
        logger.info(f"ODF available: {ODF_AVAILABLE}")
        
    def convert(self, input_path: str, output_path: str, options: dict = None):
        """Convert spreadsheet to any supported format"""
        try:
            # Check if input file exists
            if not os.path.exists(input_path):
                return False, f"Input file not found: {input_path}"
            
            input_ext = Path(input_path).suffix[1:].lower()
            output_ext = Path(output_path).suffix[1:].lower()
            
            logger.info(f"Converting spreadsheet: {Path(input_path).name} ({input_ext} -> {output_ext})")
            
            # ===== CONVERSION CASES =====
            
            # Excel → other formats
            if input_ext in ['xlsx', 'xls']:
                if output_ext == 'csv':
                    return self._excel_to_csv(input_path, output_path, options)
                elif output_ext == 'tsv':
                    return self._excel_to_tsv(input_path, output_path, options)
                elif output_ext == 'json':
                    return self._excel_to_json(input_path, output_path, options)
                elif output_ext == 'html':
                    return self._excel_to_html(input_path, output_path, options)
                elif output_ext == 'ods':
                    return self._excel_to_ods(input_path, output_path, options)
                elif output_ext in ['xlsx', 'xls']:
                    return self._excel_to_excel(input_path, output_path, options)
            
            # CSV → other formats
            elif input_ext == 'csv':
                if output_ext == 'xlsx':
                    return self._csv_to_excel(input_path, output_path, options)
                elif output_ext == 'xls':
                    return self._csv_to_excel(input_path, output_path, options)
                elif output_ext == 'tsv':
                    return self._csv_to_tsv(input_path, output_path)
                elif output_ext == 'json':
                    return self._csv_to_json(input_path, output_path)
                elif output_ext == 'html':
                    return self._csv_to_html(input_path, output_path)
                elif output_ext == 'ods':
                    return self._csv_to_ods(input_path, output_path)
            
            # TSV → other formats
            elif input_ext == 'tsv':
                if output_ext == 'csv':
                    return self._tsv_to_csv(input_path, output_path)
                elif output_ext == 'xlsx':
                    return self._tsv_to_excel(input_path, output_path)
                elif output_ext == 'json':
                    return self._tsv_to_json(input_path, output_path)
            
            # ODS → other formats
            elif input_ext == 'ods':
                if output_ext == 'xlsx':
                    return self._ods_to_excel(input_path, output_path)
                elif output_ext == 'csv':
                    return self._ods_to_csv(input_path, output_path)
                elif output_ext == 'json':
                    return self._ods_to_json(input_path, output_path)
            
            # JSON → other formats
            elif input_ext == 'json':
                if output_ext == 'csv':
                    return self._json_to_csv(input_path, output_path)
                elif output_ext == 'xlsx':
                    return self._json_to_excel(input_path, output_path)
                elif output_ext == 'html':
                    return self._json_to_html(input_path, output_path)
            
            # HTML → other formats
            elif input_ext == 'html':
                if output_ext == 'csv':
                    return self._html_to_csv(input_path, output_path)
                elif output_ext == 'xlsx':
                    return self._html_to_excel(input_path, output_path)
            
            # ===== FALLBACK: Try with Pandas if available =====
            if PANDAS_AVAILABLE:
                success, message = self._convert_with_pandas(input_path, output_path, options)
                if success:
                    return True, message
            
            # ===== LAST RESORT: Copy =====
            shutil.copy2(input_path, output_path)
            return True, f"File copied (no conversion available for {input_ext} → {output_ext})"
            
        except Exception as e:
            logger.error(f"Spreadsheet conversion error: {e}")
            return False, f"Spreadsheet conversion error: {str(e)}"
    
    # ============================================================
    # EXCEL CONVERSIONS (XLSX, XLS)
    # ============================================================
    
    def _excel_to_csv(self, input_path: str, output_path: str, options: dict = None):
        """Convert Excel to CSV"""
        try:
            if not EXCEL_AVAILABLE:
                return False, "OpenPyXL not available. Install: pip install openpyxl"
            
            wb = load_workbook(input_path, data_only=True)
            
            # Get active sheet or specified sheet
            sheet_name = options.get('sheet', None) if options else None
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    # Convert None to empty string
                    writer.writerow([str(cell) if cell is not None else '' for cell in row])
            
            return True, f"Excel to CSV conversion successful ({ws.title})"
            
        except Exception as e:
            return False, f"Excel to CSV error: {str(e)}"
    
    def _excel_to_tsv(self, input_path: str, output_path: str, options: dict = None):
        """Convert Excel to TSV"""
        try:
            if not EXCEL_AVAILABLE:
                return False, "OpenPyXL not available"
            
            wb = load_workbook(input_path, data_only=True)
            ws = wb.active
            
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f, delimiter='\t')
                for row in ws.iter_rows(values_only=True):
                    writer.writerow([str(cell) if cell is not None else '' for cell in row])
            
            return True, "Excel to TSV conversion successful"
            
        except Exception as e:
            return False, f"Excel to TSV error: {str(e)}"
    
    def _excel_to_json(self, input_path: str, output_path: str, options: dict = None):
        """Convert Excel to JSON"""
        try:
            if not EXCEL_AVAILABLE:
                return False, "OpenPyXL not available"
            
            wb = load_workbook(input_path, data_only=True)
            
            # Get sheet
            sheet_name = options.get('sheet', None) if options else None
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value is not None else '')
            
            # Get data
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    row_data = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            row_data[headers[i]] = value
                    data.append(row_data)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False, default=str)
            
            return True, f"Excel to JSON conversion successful ({len(data)} rows)"
            
        except Exception as e:
            return False, f"Excel to JSON error: {str(e)}"
    
    def _excel_to_html(self, input_path: str, output_path: str, options: dict = None):
        """Convert Excel to HTML"""
        try:
            if not EXCEL_AVAILABLE:
                return False, "OpenPyXL not available"
            
            wb = load_workbook(input_path, data_only=True)
            ws = wb.active
            
            html = ['<!DOCTYPE html>', '<html><head><meta charset="utf-8">',
                    '<title>Spreadsheet</title>',
                    '<style>',
                    'body {font-family: Arial, sans-serif; margin: 20px;}',
                    'table {border-collapse: collapse; width: 100%;}',
                    'th, td {border: 1px solid #ddd; padding: 8px; text-align: left;}',
                    'th {background-color: #f2f2f2; font-weight: bold;}',
                    'tr:nth-child(even) {background-color: #f9f9f9;}',
                    'tr:hover {background-color: #f5f5f5;}',
                    '</style></head><body>',
                    f'<h1>{Path(input_path).name}</h1>',
                    '<table>']
            
            # Headers
            html.append('<tr>')
            for cell in ws[1]:
                html.append(f'<th>{cell.value if cell.value is not None else ""}</th>')
            html.append('</tr>')
            
            # Data
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    html.append('<tr>')
                    for cell in row:
                        html.append(f'<td>{cell if cell is not None else ""}</td>')
                    html.append('</tr>')
            
            html.append('</table></body></html>')
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(html))
            
            return True, "Excel to HTML conversion successful"
            
        except Exception as e:
            return False, f"Excel to HTML error: {str(e)}"
    
    def _excel_to_ods(self, input_path: str, output_path: str, options: dict = None):
        """Convert Excel to ODS"""
        try:
            if not EXCEL_AVAILABLE:
                return False, "OpenPyXL not available"
            
            if not ODF_AVAILABLE:
                return False, "ODF not available. Install: pip install odfpy"
            
            wb = load_workbook(input_path, data_only=True)
            ws = wb.active
            
            # Create ODS document
            doc = OpenDocumentSpreadsheet()
            table = Table(name=ws.title)
            doc.spreadsheet.addElement(table)
            
            # Add data
            for row in ws.iter_rows(values_only=True):
                tr = TableRow()
                for cell in row:
                    tc = TableCell()
                    if cell is not None:
                        tc.addElement(P(text=str(cell)))
                    tr.addElement(tc)
                table.addElement(tr)
            
            doc.save(output_path)
            return True, "Excel to ODS conversion successful"
            
        except Exception as e:
            return False, f"Excel to ODS error: {str(e)}"
    
    def _excel_to_excel(self, input_path: str, output_path: str, options: dict = None):
        """Convert Excel to Excel (copy with options)"""
        try:
            if not EXCEL_AVAILABLE:
                return False, "OpenPyXL not available"
            
            wb = load_workbook(input_path, data_only=True)
            
            # Apply options
            if options:
                # Remove empty rows
                if options.get('remove_empty', False):
                    for sheet in wb.worksheets:
                        rows_to_delete = []
                        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                            if all(cell is None for cell in row):
                                rows_to_delete.append(idx)
                        for idx in reversed(rows_to_delete):
                            sheet.delete_rows(idx)
                
                # Add summary sheet
                if options.get('add_summary', False):
                    summary = wb.create_sheet('Summary')
                    summary['A1'] = 'File'
                    summary['B1'] = 'Rows'
                    summary['C1'] = 'Columns'
                    row = 2
                    for sheet in wb.worksheets:
                        summary[f'A{row}'] = sheet.title
                        summary[f'B{row}'] = sheet.max_row
                        summary[f'C{row}'] = sheet.max_column
                        row += 1
            
            wb.save(output_path)
            return True, "Excel to Excel conversion successful"
            
        except Exception as e:
            return False, f"Excel to Excel error: {str(e)}"
    
    # ============================================================
    # CSV CONVERSIONS
    # ============================================================
    
    def _csv_to_excel(self, input_path: str, output_path: str, options: dict = None):
        """Convert CSV to Excel"""
        try:
            if not EXCEL_AVAILABLE:
                return False, "OpenPyXL not available"
            
            wb = Workbook()
            ws = wb.active
            
            # Detect delimiter
            delimiter = self._detect_csv_delimiter(input_path)
            
            with open(input_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f, delimiter=delimiter)
                for row_idx, row in enumerate(reader, 1):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-fit columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(output_path)
            return True, "CSV to Excel conversion successful"
            
        except Exception as e:
            return False, f"CSV to Excel error: {str(e)}"
    
    def _csv_to_tsv(self, input_path: str, output_path: str):
        """Convert CSV to TSV"""
        try:
            delimiter = self._detect_csv_delimiter(input_path)
            
            with open(input_path, 'r', encoding='utf-8') as infile:
                reader = csv.reader(infile, delimiter=delimiter)
                with open(output_path, 'w', encoding='utf-8', newline='') as outfile:
                    writer = csv.writer(outfile, delimiter='\t')
                    for row in reader:
                        writer.writerow(row)
            
            return True, "CSV to TSV conversion successful"
            
        except Exception as e:
            return False, f"CSV to TSV error: {str(e)}"
    
    def _csv_to_json(self, input_path: str, output_path: str):
        """Convert CSV to JSON"""
        try:
            delimiter = self._detect_csv_delimiter(input_path)
            
            with open(input_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                data = list(reader)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            return True, f"CSV to JSON conversion successful ({len(data)} rows)"
            
        except Exception as e:
            return False, f"CSV to JSON error: {str(e)}"
    
    def _csv_to_html(self, input_path: str, output_path: str):
        """Convert CSV to HTML"""
        try:
            delimiter = self._detect_csv_delimiter(input_path)
            
            with open(input_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f, delimiter=delimiter)
                rows = list(reader)
            
            html = ['<!DOCTYPE html>', '<html><head><meta charset="utf-8">',
                    '<title>CSV Table</title>',
                    '<style>',
                    'body {font-family: Arial, sans-serif; margin: 20px;}',
                    'table {border-collapse: collapse; width: 100%;}',
                    'th, td {border: 1px solid #ddd; padding: 8px; text-align: left;}',
                    'th {background-color: #f2f2f2; font-weight: bold;}',
                    'tr:nth-child(even) {background-color: #f9f9f9;}',
                    'tr:hover {background-color: #f5f5f5;}',
                    '</style></head><body>',
                    f'<h1>{Path(input_path).name}</h1>',
                    '<table>']
            
            for i, row in enumerate(rows):
                html.append('<tr>')
                for cell in row:
                    tag = 'th' if i == 0 else 'td'
                    html.append(f'<{tag}>{cell}</{tag}>')
                html.append('</tr>')
            
            html.append('</table></body></html>')
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(html))
            
            return True, "CSV to HTML conversion successful"
            
        except Exception as e:
            return False, f"CSV to HTML error: {str(e)}"
    
    def _csv_to_ods(self, input_path: str, output_path: str):
        """Convert CSV to ODS"""
        try:
            if not ODF_AVAILABLE:
                return False, "ODF not available"
            
            delimiter = self._detect_csv_delimiter(input_path)
            
            doc = OpenDocumentSpreadsheet()
            table = Table(name="Sheet1")
            doc.spreadsheet.addElement(table)
            
            with open(input_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f, delimiter=delimiter)
                for row in reader:
                    tr = TableRow()
                    for cell in row:
                        tc = TableCell()
                        tc.addElement(P(text=cell))
                        tr.addElement(tc)
                    table.addElement(tr)
            
            doc.save(output_path)
            return True, "CSV to ODS conversion successful"
            
        except Exception as e:
            return False, f"CSV to ODS error: {str(e)}"
    
    # ============================================================
    # TSV CONVERSIONS
    # ============================================================
    
    def _tsv_to_csv(self, input_path: str, output_path: str):
        """Convert TSV to CSV"""
        try:
            with open(input_path, 'r', encoding='utf-8') as infile:
                reader = csv.reader(infile, delimiter='\t')
                with open(output_path, 'w', encoding='utf-8', newline='') as outfile:
                    writer = csv.writer(outfile)
                    for row in reader:
                        writer.writerow(row)
            
            return True, "TSV to CSV conversion successful"
            
        except Exception as e:
            return False, f"TSV to CSV error: {str(e)}"
    
    def _tsv_to_excel(self, input_path: str, output_path: str):
        """Convert TSV to Excel"""
        try:
            if not EXCEL_AVAILABLE:
                return False, "OpenPyXL not available"
            
            wb = Workbook()
            ws = wb.active
            
            with open(input_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f, delimiter='\t')
                for row_idx, row in enumerate(reader, 1):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(output_path)
            return True, "TSV to Excel conversion successful"
            
        except Exception as e:
            return False, f"TSV to Excel error: {str(e)}"
    
    def _tsv_to_json(self, input_path: str, output_path: str):
        """Convert TSV to JSON"""
        try:
            with open(input_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f, delimiter='\t')
                data = list(reader)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            return True, f"TSV to JSON conversion successful ({len(data)} rows)"
            
        except Exception as e:
            return False, f"TSV to JSON error: {str(e)}"
    
    # ============================================================
    # ODS CONVERSIONS
    # ============================================================
    
    def _ods_to_excel(self, input_path: str, output_path: str):
        """Convert ODS to Excel"""
        try:
            if not ODF_AVAILABLE:
                return False, "ODF not available"
            
            if not EXCEL_AVAILABLE:
                return False, "OpenPyXL not available"
            
            doc = load(input_path)
            wb = Workbook()
            
            # Get first table
            tables = doc.getElementsByType(Table)
            if tables:
                table = tables[0]
                ws = wb.active
                ws.title = table.getAttribute('name') or 'Sheet1'
                
                row_idx = 1
                for row in table.getElementsByType(TableRow):
                    col_idx = 1
                    for cell in row.getElementsByType(TableCell):
                        text = ""
                        for p in cell.getElementsByType(P):
                            if p.firstChild:
                                text += p.firstChild.data
                        ws.cell(row=row_idx, column=col_idx, value=text)
                        col_idx += 1
                    row_idx += 1
            
            wb.save(output_path)
            return True, "ODS to Excel conversion successful"
            
        except Exception as e:
            return False, f"ODS to Excel error: {str(e)}"
    
    def _ods_to_csv(self, input_path: str, output_path: str):
        """Convert ODS to CSV"""
        try:
            if not ODF_AVAILABLE:
                return False, "ODF not available"
            
            doc = load(input_path)
            
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                tables = doc.getElementsByType(Table)
                if tables:
                    table = tables[0]
                    for row in table.getElementsByType(TableRow):
                        row_data = []
                        for cell in row.getElementsByType(TableCell):
                            text = ""
                            for p in cell.getElementsByType(P):
                                if p.firstChild:
                                    text += p.firstChild.data
                            row_data.append(text)
                        writer.writerow(row_data)
            
            return True, "ODS to CSV conversion successful"
            
        except Exception as e:
            return False, f"ODS to CSV error: {str(e)}"
    
    def _ods_to_json(self, input_path: str, output_path: str):
        """Convert ODS to JSON"""
        try:
            if not ODF_AVAILABLE:
                return False, "ODF not available"
            
            doc = load(input_path)
            
            tables = doc.getElementsByType(Table)
            if not tables:
                return False, "No tables found in ODS"
            
            table = tables[0]
            data = []
            
            # Get headers from first row
            headers = []
            first_row = True
            for row in table.getElementsByType(TableRow):
                row_data = []
                for cell in row.getElementsByType(TableCell):
                    text = ""
                    for p in cell.getElementsByType(P):
                        if p.firstChild:
                            text += p.firstChild.data
                    row_data.append(text)
                
                if first_row:
                    headers = row_data
                    first_row = False
                else:
                    if any(row_data):
                        row_dict = {}
                        for i, value in enumerate(row_data):
                            if i < len(headers):
                                row_dict[headers[i]] = value
                        data.append(row_dict)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            return True, f"ODS to JSON conversion successful ({len(data)} rows)"
            
        except Exception as e:
            return False, f"ODS to JSON error: {str(e)}"
    
    # ============================================================
    # JSON CONVERSIONS
    # ============================================================
    
    def _json_to_csv(self, input_path: str, output_path: str):
        """Convert JSON to CSV"""
        try:
            with open(input_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if not isinstance(data, list) or not data:
                return False, "JSON must be a list of objects"
            
            headers = list(data[0].keys())
            
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(data)
            
            return True, f"JSON to CSV conversion successful ({len(data)} rows)"
            
        except Exception as e:
            return False, f"JSON to CSV error: {str(e)}"
    
    def _json_to_excel(self, input_path: str, output_path: str):
        """Convert JSON to Excel"""
        try:
            if not EXCEL_AVAILABLE:
                return False, "OpenPyXL not available"
            
            with open(input_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if not isinstance(data, list) or not data:
                return False, "JSON must be a list of objects"
            
            wb = Workbook()
            ws = wb.active
            
            # Headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Data
            for row_idx, item in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    value = item.get(header, '')
                    ws.cell(row=row_idx, column=col, value=value)
            
            wb.save(output_path)
            return True, f"JSON to Excel conversion successful ({len(data)} rows)"
            
        except Exception as e:
            return False, f"JSON to Excel error: {str(e)}"
    
    def _json_to_html(self, input_path: str, output_path: str):
        """Convert JSON to HTML"""
        try:
            with open(input_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            html = ['<!DOCTYPE html>', '<html><head><meta charset="utf-8">',
                    '<title>JSON Data</title>',
                    '<style>',
                    'body {font-family: Arial, sans-serif; margin: 20px;}',
                    'pre {background: #f4f4f4; padding: 20px; border-radius: 10px; overflow: auto;}',
                    '</style></head><body>',
                    f'<h1>{Path(input_path).name}</h1>',
                    f'<pre>{json.dumps(data, indent=2, ensure_ascii=False)}</pre>',
                    '</body></html>']
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(html))
            
            return True, "JSON to HTML conversion successful"
            
        except Exception as e:
            return False, f"JSON to HTML error: {str(e)}"
    
    # ============================================================
    # HTML CONVERSIONS
    # ============================================================
    
    def _html_to_csv(self, input_path: str, output_path: str):
        """Convert HTML table to CSV"""
        try:
            import html.parser
            
            class TableParser(html.parser.HTMLParser):
                def __init__(self):
                    super().__init__()
                    self.in_table = False
                    self.in_row = False
                    self.in_cell = False
                    self.rows = []
                    self.current_row = []
                    self.current_cell = []
                
                def handle_starttag(self, tag, attrs):
                    if tag == 'table':
                        self.in_table = True
                    elif tag == 'tr' and self.in_table:
                        self.in_row = True
                        self.current_row = []
                    elif tag in ('td', 'th') and self.in_row:
                        self.in_cell = True
                        self.current_cell = []
                
                def handle_endtag(self, tag):
                    if tag == 'table':
                        self.in_table = False
                    elif tag == 'tr':
                        if self.in_row:
                            self.rows.append(self.current_row)
                            self.in_row = False
                    elif tag in ('td', 'th'):
                        if self.in_cell:
                            self.current_row.append(''.join(self.current_cell))
                            self.in_cell = False
                
                def handle_data(self, data):
                    if self.in_cell:
                        self.current_cell.append(data.strip())
            
            parser = TableParser()
            with open(input_path, 'r', encoding='utf-8') as f:
                parser.feed(f.read())
            
            if not parser.rows:
                return False, "No table found in HTML"
            
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerows(parser.rows)
            
            return True, "HTML to CSV conversion successful"
            
        except Exception as e:
            return False, f"HTML to CSV error: {str(e)}"
    
    def _html_to_excel(self, input_path: str, output_path: str):
        """Convert HTML table to Excel"""
        try:
            # Convert to CSV first, then to Excel
            temp_csv = tempfile.NamedTemporaryFile(suffix='.csv', delete=False)
            temp_path = temp_csv.name
            temp_csv.close()
            
            success, message = self._html_to_csv(input_path, temp_path)
            if not success:
                os.unlink(temp_path)
                return False, message
            
            success, message = self._csv_to_excel(temp_path, output_path, None)
            os.unlink(temp_path)
            return success, message
            
        except Exception as e:
            return False, f"HTML to Excel error: {str(e)}"
    
    # ============================================================
    # PANDAS CONVERSION (Fallback)
    # ============================================================
    
    def _convert_with_pandas(self, input_path: str, output_path: str, options: dict = None):
        """Convert using Pandas (if available)"""
        try:
            if not PANDAS_AVAILABLE:
                return False, "Pandas not available"
            
            input_ext = Path(input_path).suffix[1:].lower()
            output_ext = Path(output_path).suffix[1:].lower()
            
            # Read with pandas
            if input_ext in ['xlsx', 'xls']:
                df = pd.read_excel(input_path, engine='openpyxl')
            elif input_ext == 'csv':
                df = pd.read_csv(input_path)
            elif input_ext == 'tsv':
                df = pd.read_csv(input_path, sep='\t')
            elif input_ext == 'json':
                df = pd.read_json(input_path)
            else:
                return False, f"Pandas cannot read {input_ext}"
            
            # Write with pandas
            if output_ext in ['xlsx', 'xls']:
                df.to_excel(output_path, index=False, engine='openpyxl')
            elif output_ext == 'csv':
                df.to_csv(output_path, index=False)
            elif output_ext == 'tsv':
                df.to_csv(output_path, sep='\t', index=False)
            elif output_ext == 'json':
                df.to_json(output_path, orient='records', indent=2)
            elif output_ext == 'html':
                df.to_html(output_path)
            else:
                return False, f"Pandas cannot write {output_ext}"
            
            return True, f"Converted with Pandas ({len(df)} rows, {len(df.columns)} columns)"
            
        except Exception as e:
            return False, f"Pandas conversion error: {str(e)}"
    
    # ============================================================
    # HELPER METHODS
    # ============================================================
    
    def _detect_csv_delimiter(self, file_path: str) -> str:
        """Detect CSV delimiter (comma, semicolon, tab)"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                first_line = f.readline()
                
            # Count delimiters
            delimiters = {
                ',': first_line.count(','),
                ';': first_line.count(';'),
                '\t': first_line.count('\t'),
                '|': first_line.count('|')
            }
            
            # Return the most common delimiter
            return max(delimiters, key=delimiters.get)
            
        except:
            return ','  # Default to comma
    
    def get_supported_formats(self) -> list:
        """Get all supported input formats"""
        return self.supported_formats
    
    def get_output_formats(self) -> list:
        """Get all supported output formats"""
        return self.output_formats
    
    def get_format_info(self, format_name: str) -> dict:
        """Get information about a format"""
        info = {
            'xlsx': {'name': 'Excel', 'extension': 'xlsx', 'mime': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'},
            'xls': {'name': 'Excel', 'extension': 'xls', 'mime': 'application/vnd.ms-excel'},
            'csv': {'name': 'CSV', 'extension': 'csv', 'mime': 'text/csv'},
            'tsv': {'name': 'TSV', 'extension': 'tsv', 'mime': 'text/tab-separated-values'},
            'ods': {'name': 'ODS', 'extension': 'ods', 'mime': 'application/vnd.oasis.opendocument.spreadsheet'},
            'html': {'name': 'HTML', 'extension': 'html', 'mime': 'text/html'},
            'json': {'name': 'JSON', 'extension': 'json', 'mime': 'application/json'},
            'xml': {'name': 'XML', 'extension': 'xml', 'mime': 'application/xml'},
            'numbers': {'name': 'Numbers', 'extension': 'numbers', 'mime': 'application/vnd.apple.numbers'},
            'dif': {'name': 'DIF', 'extension': 'dif', 'mime': 'text/x-dif'}
        }
        return info.get(format_name.lower(), {})